from bs4 import BeautifulSoup as bs
import requests
import yagmail
from openpyxl import load_workbook

wb = load_workbook("flipkart_tracker.xlsx")
ws = wb.active
row_count = ws.max_row

print("Processing...........")


def send_alert(reciever):
    for col in range(2, row_count + 1):
        Link = ws[f"B{col}"].value
        Pricing = ws[f"C{col}"].value
        Name = ws[f"A{col}"].value.lstrip()

        response = requests.get(Link, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36",
            "Accept-Language": "en-US,en;q=0.9"})

        html = response.text
        data = bs(html, "html.parser")
        name = data.find("span", class_="B_NuCI").get_text().lstrip()
        latest_price = float(data.find("div", class_="_30jeq3 _16Jk6d").get_text().split("₹")[1].replace(",", ""))


        if (latest_price < Pricing):
            subject = f"Price Drop Alert: {Name}"
            contents = [
                f"We are pleased to inform you that the price of '{name}' on Flipkart has dropped to a new low of Rs,{latest_price}. This price reduction is below your desired threshold of Rs{Pricing}, making it an excellent time to consider your purchase.\n\n",

                f"Product Details:\n\n",
                f"Product Name: {name}\n\n",
                f"Current Price: Rs,{latest_price}\n",
                f"Original Price: Rs,{Pricing}\n\n",
                f"Product Link: {Link}\n\n",

                f"Hurry and seize this opportunity to make your purchase while the price is favorable.\n\n",

                f"Thank you for using our Flipkart price tracking service",

            ]
            contents = ''.join(contents)

            yagmail.SMTP('Email Username', 'Email Password').send(reciever, subject, contents)
            print(f"Price Drop Detected! . Succesfully Send The Alert Message To {reciever}")
        ws[f"C{col}"].value = latest_price


if __name__ == "__main__":

    reciever = "ashharjosh@gmail.com"

    try:
        send_alert(reciever)
    except:
        print("Error Tracking The Prices")

    wb.save("flipkart_tracker.xlsx")
    print("Finished Checking For Price Alert")